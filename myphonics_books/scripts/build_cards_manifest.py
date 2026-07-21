"""
MyPhonicsBooks — Sound Cards manifest builder.

Regenerates output/worksheet_plan/Sound_Cards_Manifest.xlsx from
output/worksheet_plan/sound_cards.json so the manifest can never drift
from the card data again (it was originally a one-off export; when the
`sc` card was added on 2026-07-07 there was no script to refresh it).

Two sheets, mirroring the original workbook:
  Summary        — per-level Main/Twin/Extra counts vs the build-spec
                   estimates, plus the built physical-deck total
                   (mains + one shifty per twin group + extras).
  Card Manifest  — one row per raw card + the ough insert rows.

Run:  py -3.12 scripts/build_cards_manifest.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR / "scripts"))

import cards_data as cd  # noqa: E402

OUT_XLSX = BASE_DIR / "output" / "worksheet_plan" / "Sound_Cards_Manifest.xlsx"

# Build-spec estimates from the 2026-07-04 spec (historical reference —
# do not update these when the deck changes; the delta is the point).
SPEC_ESTIMATES = {
    "L1": "10", "L2": "19", "L3": "9", "L4": "12",
    "L5": "17", "L6": "24", "L7": "26", "L8": "18",
    "Total": "136 spec",
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
BODY_FONT = Font(size=10, color="111827")
BODY_BOLD = Font(bold=True, size=10, color="111827")
LEVEL_FONT = Font(bold=True, size=10, color="FFFFFF")
TOTAL_FONT = Font(bold=True, size=10, color="166534")
TOTAL_FILL = PatternFill("solid", fgColor="DCFCE7")
TITLE_FONT = Font(bold=True, size=15, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1A1A1A")

TYPE_FILLS = {
    "main": "FAF7F1",
    "twin": "E8F3EC",
    "extra_spelling": "F2EBDD",
    "wildcard": "F2EBDD",
    "wildcard_title": "F2EBDD",
}


def build() -> None:
    raw = cd.load_raw()
    cards = raw["cards"]
    ough = raw["ough_insert_set"]
    deck, _ = cd.load_cards("all")

    n_main = sum(1 for c in cards if c["type"] == "main")
    n_twin = sum(1 for c in cards if c["type"] == "twin")
    n_extra = sum(1 for c in cards if c["type"] == "extra_spelling")
    n_shifty = sum(1 for c in deck if c["kind"] == "shifty")

    wb = Workbook()

    # ── Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    for col, width in zip("ABCDEF", (12, 16, 16, 16, 16, 16)):
        ws.column_dimensions[col].width = width
    ws["A1"] = "Sound Cards — Manifest & Counts"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL

    headers = ("Level", "Main (A)", "Twin (B)", "Extra (C)",
               "Level total", "Spec estimate")
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL

    level_colours = {c["level"]: c["level_colour"] for c in cards}
    row = 4
    for lv in cd.LEVEL_ORDER:
        counts = [
            sum(1 for c in cards if c["level"] == lv and c["type"] == t)
            for t in ("main", "twin", "extra_spelling")
        ]
        lc = ws.cell(row=row, column=1, value=lv)
        lc.font = LEVEL_FONT
        lc.fill = PatternFill("solid", fgColor=level_colours[lv].lstrip("#"))
        for i, v in enumerate([*counts, sum(counts), SPEC_ESTIMATES[lv]], start=2):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = BODY_FONT
        row += 1

    totals = ("Total", n_main, n_twin, n_extra, len(cards),
              SPEC_ESTIMATES["Total"])
    for i, v in enumerate(totals, start=1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font, cell.fill = TOTAL_FONT, TOTAL_FILL

    note = (f"Raw scheme: {len(cards)} cards ({n_main} main + {n_twin} twin + "
            f"{n_extra} extra). Built physical deck: {len(deck)} cards "
            f"({n_main} main + {n_shifty} shifty + {n_extra} extra — twins fold "
            f"onto one shifty card per grapheme). Plus ough insert set: "
            f"{len(ough)} cards. Grand total printed: {len(deck) + len(ough)}.")
    nc = ws.cell(row=row + 2, column=1, value=note)
    nc.font = BODY_BOLD

    # ── Card Manifest ────────────────────────────────────────────
    ws2 = wb.create_sheet("Card Manifest")
    for col, width in zip("ABCDEFGHI", (20, 14, 7, 10, 16, 12, 10, 7, 50)):
        ws2.column_dimensions[col].width = width
    ws2.freeze_panes = "A2"
    headers2 = ("Card ID", "Type", "Level", "Grapheme", "Says", "Key word",
                "Twin", "Words", "First words")
    for i, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL

    def write_card(r: int, c: dict) -> None:
        words = c.get("words") or []
        twin = (f"{c['twin_number']}/{c['twin_total']}"
                if c.get("twin_number") else None)
        first_words = ", ".join(words[:10]) if words else c.get("note")
        values = (c["card_id"], c["type"], c["level"], c.get("grapheme"),
                  c.get("says_plain"), c.get("key_word"), twin,
                  len(words) or None, first_words)
        for i, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=i, value=v)
            cell.font = BODY_FONT
        ws2.cell(row=r, column=2).fill = PatternFill(
            "solid", fgColor=TYPE_FILLS.get(c["type"], "FFFFFF"))
        lvl = ws2.cell(row=r, column=3)
        lvl.font = LEVEL_FONT
        lvl.fill = PatternFill("solid", fgColor=c["level_colour"].lstrip("#"))
        ws2.cell(row=r, column=4).font = BODY_BOLD

    r = 2
    for c in cards + ough:
        write_card(r, c)
        r += 1

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX.relative_to(BASE_DIR)} — "
          f"{len(cards)} raw cards, physical deck {len(deck)}, "
          f"ough insert {len(ough)} rows")


if __name__ == "__main__":
    build()

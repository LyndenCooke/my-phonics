"""Printable word-card decks from the Sound Scheme workbook.

Reads output/worksheet_plan/MyPhonicsBooks_Sound_Scheme_CORRECTED.xlsx and
renders double-sided A4 card sheets (2 cols x 4 rows = 8 cards per side):

  - Cell 1 of every sheet is the SOUND CARD (grapheme front / grapheme +
    says + key word back), cells 2-8 are that sound's practice words.
  - Page 1 of each pair = plain words (child-facing side).
  - Page 2 = the same words with decoding marks, columns MIRRORED so that
    long-edge double-sided printing puts the marked word exactly behind
    its plain twin. Cut along the grey grid lines.

Decoding-mark grammar (locked with Lynden 2026-07-05):
  dot        = one letter, one sound
  short line = these letters together make one sound (straight, NOT curved)
  over-arc   = split digraph (curved, over the top, vowel to silent e)
  diamond    = this grapheme is NOT making its usual taught sound (slate,
               letters tinted slate) - alt pronunciations only; extra
               spellings / late-taught GPCs get normal marks.

A word is diamond-marked wherever it appears: every word is cross-checked
against ALL 'alt pronunciation' rows, so "was" gets its a-diamond and
s-diamond even on another sound's card.

Run:  py -3.12 scripts/generate_word_cards.py            # everything
      py -3.12 scripts/generate_word_cards.py --level 6  # one level
      py -3.12 scripts/generate_word_cards.py --dry-run  # marks to stdout
Output: output/word_cards/L{n}_sound_cards.pdf + L{n}_shifty_cards.pdf
"""

import argparse
import asyncio
import base64
import json
import re
import sys
from pathlib import Path

import openpyxl
from jinja2 import Environment, FileSystemLoader

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR / "scripts"))

import v2_helpers
from v2_helpers import _compute_marks

XLSX = BASE_DIR / "output" / "worksheet_plan" / "MyPhonicsBooks_Sound_Scheme_CORRECTED.xlsx"
FONTS_DIR = BASE_DIR / "assets" / "fonts"
OUT_DIR = BASE_DIR / "output" / "word_cards"
BUILD_DIR = OUT_DIR / "_build"

LEVEL_COLOURS = {
    1: "#E84B8A", 2: "#F97066", 3: "#F59E0B", 4: "#22C55E",
    5: "#3B82F6", 6: "#6366F1", 7: "#8B5CF6", 8: "#14B8A6",
}
LEVEL_NAMES = {
    1: "Ditties", 2: "First Sounds", 3: "Special Friends", 4: "Longer Sounds",
    5: "New Spellings", 6: "Building Fluency", 7: "Reading Together", 8: "Reading Champion",
}
SHIFTY_COLOUR = "#475569"
WORDS_PER_SHEET = 7  # cell 1 is the sound card

# Graphemes safe to match in ANY word (unambiguous letter runs).  Riskier
# shifty graphemes (se, ve, ce, ge, ui, oe, ey, au...) are added per-card
# only, so "sea" never splits as se+a on an unrelated card.
BASE_EXTRA_GRAPHEMES = [
    "tch", "dge", "ough", "augh", "eigh", "wh", "ph", "kn", "wr", "mb", "gn",
]
# Cards whose focus grapheme swallows the final e — the magic-e detector
# must stay OFF for their words (have is ve, not a-e + v).
NO_SPLIT_FOCUS = {"ve", "se", "ce", "ge"}


def load_base_graphemes() -> list:
    data = json.loads((BASE_DIR / "data" / "graphemes_by_level.json").read_text(encoding="utf-8"))
    out = []
    for lv in data.values():
        out.extend(lv["graphemes"])
    out.extend(sorted(v2_helpers.DOUBLED_CONSONANTS))
    out.extend(BASE_EXTRA_GRAPHEMES)
    return out


def clean_words(raw: str) -> list:
    """Split the sheet's comma list, strip '(gloss)' notes, dedupe in order."""
    words, seen = [], set()
    for w in (raw or "").split(","):
        w = re.sub(r"\([^)]*\)", "", w).strip()
        if w and w.lower() not in seen:
            seen.add(w.lower())
            words.append(w)
    return words


def read_workbook():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    sound_cards = []
    block = level = None
    for row in wb["2. Sound Scheme"].iter_rows(min_row=4, values_only=True):
        blk, grapheme, says, key, words, _count = row
        if blk:
            block = blk.split("\n")[0].strip()
            level = int(re.search(r"L(\d)", blk).group(1))
        if not grapheme:
            continue
        sound_cards.append({
            "kind": "sound", "level": level, "block": block,
            "grapheme": str(grapheme).strip(), "says": (says or "").strip(),
            "key": (key or "").strip(), "words": clean_words(words),
            "diamond": False,
        })

    shifty_cards = []
    for row in wb["3. Shifty Sounds"].iter_rows(min_row=4, values_only=True):
        frm, typ, grapheme, says, words, _count = row
        if not grapheme:
            continue
        level = int(re.search(r"L(\d)", str(frm)).group(1))
        typ = (typ or "").strip()
        shifty_cards.append({
            "kind": "shifty", "level": level, "block": typ,
            "grapheme": str(grapheme).strip(), "says": (says or "").strip(),
            "key": "", "words": clean_words(words),
            "diamond": typ == "alt pronunciation",
        })

    return sound_cards, shifty_cards


def build_shifty_index(shifty_cards: list) -> dict:
    """word -> [grapheme, ...] for every alt-pronunciation membership."""
    index = {}
    for card in shifty_cards:
        if not card["diamond"]:
            continue
        g = card["grapheme"].lower()
        if not g.isalpha():
            continue
        for w in card["words"]:
            index.setdefault(w.lower(), []).append(g)
    return index


def apply_diamond(marks: list, word: str, grapheme: str, focus: str = "") -> list:
    """Replace the mark covering `grapheme`'s first occurrence with a diamond.

    `focus` is the card's own grapheme: a cross-referenced diamond never
    breaks apart a matched focus unit (on the `ce` card, "ice" keeps its
    ce line rather than being split for the c=/s/ diamond).
    """
    lower = word.lower()
    focus_spans = [set(m["indices"]) for m in marks
                   if focus and focus != grapheme
                   and "".join(lower[j] for j in m["indices"]) == focus]
    # Prefer an existing mark that IS exactly this grapheme (dot or line)
    for i, m in enumerate(marks):
        if m["type"] in ("dot", "under_arc"):
            covered = "".join(lower[j] for j in m["indices"])
            if covered == grapheme:
                marks[i] = {"type": "diamond", "indices": m["indices"]}
                return marks
    # Fallback: first raw occurrence; drop anything overlapping (silent e
    # from a bogus magic-e arc simply loses its mark)
    pos = lower.find(grapheme)
    if pos < 0:
        return marks
    span = set(range(pos, pos + len(grapheme)))
    if any(span & fs for fs in focus_spans):
        return marks
    marks = [m for m in marks if not (set(m["indices"]) & span)]
    marks.append({"type": "diamond", "indices": sorted(span)})
    return sorted(marks, key=lambda m: m["indices"][0])


def force_end_unit(marks: list, word: str, suffix: str) -> list:
    """Guarantee a trailing suffix ('-ed' card) renders as one line unit."""
    lower = word.lower()
    if not lower.endswith(suffix):
        return marks
    span = set(range(len(lower) - len(suffix), len(lower)))
    for m in marks:
        if set(m["indices"]) == span:
            return marks  # already a unit
    marks = [m for m in marks if not (set(m["indices"]) & span)]
    marks.append({"type": "under_arc", "indices": sorted(span)})
    return sorted(marks, key=lambda m: m["indices"][0])


def word_segments(word: str, inventory: list, shifty_index: dict,
                  allow_split: bool = True, focus: str = "") -> list:
    """Word -> renderable segment list.

    Segments: {'kind':'unit','text','mark'} where mark is dot|line|diamond|none,
    or {'kind':'arc','units':[...]} for a split-digraph over-arc group.
    """
    saved = v2_helpers.SPLIT_DIGRAPHS
    try:
        if not allow_split:
            v2_helpers.SPLIT_DIGRAPHS = []
        marks = _compute_marks(word, inventory)
    finally:
        v2_helpers.SPLIT_DIGRAPHS = saved

    if focus.startswith("-") and focus[1:].isalpha():
        marks = force_end_unit(marks, word, focus[1:])

    for g in shifty_index.get(word.lower(), []):
        marks = apply_diamond(marks, word, g, focus)

    # Index -> mark lookups
    arc = next((m for m in marks if m["type"] == "over_arc"), None)
    arc_span = set(arc["indices"]) if arc else set()
    by_start = {}
    for m in marks:
        if m["type"] == "over_arc":
            continue
        by_start[m["indices"][0]] = m

    def unit(m, i):
        text = word[m["indices"][0]: m["indices"][-1] + 1]
        mark = {"dot": "dot", "under_arc": "line", "diamond": "diamond"}[m["type"]]
        return {"kind": "unit", "text": text, "mark": mark}

    segments, i, n = [], 0, len(word)
    while i < n:
        if arc and i == min(arc_span):
            units, j, top = [], i, max(arc_span)
            while j <= top:
                m = by_start.get(j)
                if m:
                    units.append(unit(m, j))
                    j = m["indices"][-1] + 1
                else:
                    units.append({"kind": "unit", "text": word[j], "mark": "none"})
                    j += 1
            segments.append({"kind": "arc", "units": units})
            i = top + 1
            continue
        m = by_start.get(i)
        if m:
            segments.append(unit(m, i))
            i = m["indices"][-1] + 1
        else:
            segments.append({"kind": "unit", "text": word[i], "mark": "none"})
            i += 1
    return segments


def size_class(word: str) -> str:
    n = len(word)
    if n <= 5:
        return "w-s"
    if n <= 8:
        return "w-m"
    if n <= 11:
        return "w-l"
    return "w-xl"


def build_sheets(card: dict, base_inventory: list, shifty_index: dict) -> list:
    """One card (sound + word list) -> list of sheet dicts (7 words each)."""
    g = card["grapheme"].lower()
    inventory = list(base_inventory)
    if g.isalpha() and g not in inventory:
        inventory.append(g)
    allow_split = g not in NO_SPLIT_FOCUS

    accent = SHIFTY_COLOUR if card["kind"] == "shifty" else LEVEL_COLOURS[card["level"]]
    if card["kind"] == "shifty":
        band = f"Shifty Sounds · from L{card['level']} · {card['block']}"
    else:
        band = f"Block {card['block']} · L{card['level']} {LEVEL_NAMES[card['level']]}"

    # Sound-card back: the grapheme wearing its own mark
    if "-" in g or not g.isalpha():
        g_mark = "none"
    elif card["diamond"]:
        g_mark = "diamond"
    elif len(g) > 1:
        g_mark = "line"
    else:
        g_mark = "dot"

    words = card["words"]
    chunks = [words[i:i + WORDS_PER_SHEET] for i in range(0, len(words), WORDS_PER_SHEET)] or [[]]
    sheets = []
    for idx, chunk in enumerate(chunks, 1):
        fronts, backs = [], []
        for w in chunk:
            fronts.append({"word": w, "size": size_class(w)})
            backs.append({"word": w, "size": size_class(w),
                          "segments": word_segments(w, inventory, shifty_index,
                                                    allow_split, card["grapheme"].lower())})
        while len(fronts) < WORDS_PER_SHEET:
            fronts.append(None)
            backs.append(None)
        sheets.append({
            "accent": accent, "band": band, "grapheme": card["grapheme"],
            "says": card["says"], "key": card["key"], "g_mark": g_mark,
            "level": card["level"], "kind": card["kind"],
            "sheet_idx": idx, "sheet_total": len(chunks),
            "fronts": fronts, "backs": backs,
        })
    return sheets


def font_data_uri(path: Path) -> str:
    return "data:font/truetype;base64," + base64.b64encode(path.read_bytes()).decode()


async def render_pdfs(jobs: list):
    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page()
        for html_path, pdf_path in jobs:
            await page.goto(html_path.as_uri(), wait_until="networkidle")
            await page.evaluate("document.fonts.ready")
            await asyncio.sleep(1.0)
            await page.pdf(path=str(pdf_path), width="210mm", height="297mm",
                           print_background=True,
                           margin={"top": "0", "bottom": "0", "left": "0", "right": "0"})
            # Guard against the truncated-render flake: page count must be
            # exactly 2x the sheet count baked into the HTML.
            import fitz
            expected = html_path.read_text(encoding="utf-8").count('class="sheet"')
            got = fitz.open(pdf_path).page_count
            if got != expected:
                raise RuntimeError(f"{pdf_path.name}: {got} pages, expected {expected}")
            print(f"  PDF  {pdf_path.name} ({got} pages)")
        await browser.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--level", type=int, help="only this level (1-8)")
    ap.add_argument("--dry-run", action="store_true", help="print marks, no PDFs")
    args = ap.parse_args()

    sound_cards, shifty_cards = read_workbook()
    base_inventory = load_base_graphemes()
    shifty_index = build_shifty_index(shifty_cards)

    if args.dry_run:
        # (word, card focus grapheme, focus in NO_SPLIT set) — mirrors the
        # per-card inventory each word would really be marked with.
        probe = [
            ("has", "s"), ("was", "a"), ("watch", "a"), ("author", "au"),
            ("bread", "ea"), ("book", "oo"), ("come", "o"), ("ice", "c"),
            ("ice", "ce"), ("these", "e-e"), ("played", "-ed"), ("while", "wh"),
            ("cake", "a-e"), ("duck", "ck"), ("fish", "sh"), ("the", "th"),
            ("even", "e"), ("mystery", "y"), ("through", "ough"),
            ("knee", "kn"), ("have", "ve"), ("love", "o"),
        ]
        for w, focus in probe:
            inv = list(base_inventory)
            if focus.isalpha() and focus not in inv:
                inv.append(focus)
            allow = focus not in NO_SPLIT_FOCUS
            segs = word_segments(w, inv, shifty_index, allow, focus)
            desc = " ".join(
                ("(" + " ".join(f"{u['text']}:{u['mark']}" for u in s["units"]) + ")^arc")
                if s["kind"] == "arc" else f"{s['text']}:{s['mark']}"
                for s in segs)
            print(f"  {w:10s} [{focus:4s}] {desc}")
        return

    env = Environment(loader=FileSystemLoader(str(BASE_DIR / "templates")), autoescape=False)
    template = env.get_template("word_cards.html")
    fonts = {
        "font_regular": font_data_uri(FONTS_DIR / "Andika-Regular.ttf"),
        "font_bold": font_data_uri(FONTS_DIR / "Andika-Bold.ttf"),
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    BUILD_DIR.mkdir(parents=True, exist_ok=True)

    groups = {}
    for card in sound_cards + shifty_cards:
        if args.level and card["level"] != args.level:
            continue
        key = (card["level"], card["kind"])
        groups.setdefault(key, []).extend(build_sheets(card, base_inventory, shifty_index))

    jobs = []
    for (level, kind), sheets in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1] != "sound")):
        stem = f"L{level}_{'sound' if kind == 'sound' else 'shifty'}_cards"
        html = template.render(sheets=sheets, **fonts)
        html_path = BUILD_DIR / f"{stem}.html"
        html_path.write_text(html, encoding="utf-8")
        jobs.append((html_path, OUT_DIR / f"{stem}.pdf"))
        n_words = sum(1 for s in sheets for f in s["fronts"] if f)
        print(f"L{level} {kind:6s} {len(sheets):3d} sheets ({len(sheets)*2} pages), {n_words} words")

    asyncio.run(render_pdfs(jobs))
    print(f"\nDone -> {OUT_DIR}")


if __name__ == "__main__":
    main()

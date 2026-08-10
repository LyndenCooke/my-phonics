"""Fill Bookvault's Bulk Upload Titles spreadsheet for the 33 storybooks.

Bookvault's own titleTemplate.xlsx is the authoritative schema — their public
OpenAPI spec has no definitions section, so the Product model's required fields
are NOT discoverable from the docs.  This script writes the technical columns
we can derive with certainty and leaves the commercial ones blank for Lynden.

Derived from the repo (never typed by hand):
  * ISBN + Title   — data/isbn_classroom.csv, the CONFIRMED register
  * Colour Pages   — read from the actual print master PDFs, not assumed
  * Binding/size   — Saddle Stitch, 210x148mm trim

Left BLANK on purpose (business decisions, see BLANK_COLUMNS):
  imprint, author, publisher, city, publication date, BIC/BISAC, RRPs.

Usage:
    py -3.12 -X utf8 scripts/build_bookvault_titles.py [path/to/titleTemplate.xlsx]

Writes output/bookvault/titleUpload_MyPhonicsBooks.xlsx — upload that at
Library > Bulk Upload Titles.  Nothing is sent anywhere by this script.
"""

from __future__ import annotations

import sys
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl

BASE_DIR = Path(__file__).parent.parent
MASTERS = BASE_DIR / "output" / "books" / "print_masters"
OUT_DIR = BASE_DIR / "output" / "bookvault"
DEFAULT_TEMPLATE = Path.home() / "Downloads" / "titleTemplate.xlsx"

# ── Spec choices ─────────────────────────────────────────────────────────────
# IDs come from the template's own Sheet2 lookup tables, not guessed.
BIND_ID, BIND_NAME = 5, "Saddle Stitch"          # 16/20pp stapled A5
TEXT_STOCK_ID, TEXT_STOCK = 157, "115gsm Coated"  # colour picture-book interior
COVER_STOCK_ID, COVER_STOCK = 156, "Self Cover"   # Lynden 2026-08-05.  Self
    # Cover prints the cover on the same stock as the interior, so the inside
    # front and inside back covers are ordinary printed pages — which is what
    # keeps all 33 books usable exactly as designed.  Cover Board would leave
    # those two faces blank and cost two pages of content per book.
    # WARNING: "Self Cover" is valid per the API (/ProductionLimits?BindID=5
    # returns CoverStocks [156, 5]) but is NOT in the bulk template's Sheet2
    # dropdown, so the bulk upload may reject it even though it is correct.
    # If it does, titles have to go through Add A Title instead.
LAM_ID, LAM_NAME = 2, "Matte Lamination"
LANGUAGE_ID, LANGUAGE = 1407, "English"
COUNTRY_ID, COUNTRY = 2146, "United Kingdom"
AUD_ID, AUDIENCE = 482, "Children/Juvenile"
HEIGHT_MM, WIDTH_MM = 210, 148                    # TRIM size (Height, Width)

# ── Publishing identity (Lynden 2026-08-05) ─────────────────────────────────
# Self-published: the publisher IS MyPhonicsBooks.  "Publisher" is the legal
# entity of record and "Imprint" the brand it publishes under — for a
# self-publisher both are the same name.  This is also why owning the Nielsen
# ISBNs matters: they name us as publisher, not the printer.
IMPRINT = PUBLISHER = "MyPhonicsBooks"
AUTHOR_FORENAME, AUTHOR_SURNAME = "Lynden", "Cooke"
CITY = "London"

# Subject classification — both looked up from Bookvault's own codelists, not
# guessed.  YQCR5 is the exact BIC code for synthetic-phonics reading schemes.
# JUV043000 shelves them as beginner readers: these are decodable STORIES, so
# juvenile fiction is right; LAN011000 (Phonetics & Phonics) is the adult
# language-arts shelf and would be wrong.
BIC_CODE = "YQCR5"        # Educational: English language: readers & reading
                          # schemes: Synthetic Phonics
BISAC_CODE = "JUV043000"  # JUVENILE FICTION / Readers / Beginner

# RRP mirrors the live /shop pricing (PhysicalShop.tsx): L1-L3 £5.99,
# L4-L8 £6.99.  Keyed on LEVEL, not page count — the interior page count is
# no longer a proxy for it now that the cover is split out.
def rrp_gbp(level: int) -> float:
    return 5.99 if level <= 3 else 6.99


# Bookvault's Page Count is the TEXT FILE only — the cover is a separate
# artwork file (Sizing Calculator, 2026-08-05: cover spread 303x216mm with
# 0mm spine, text block 216x154mm) and carries the two OUTSIDE faces only — confirmed by Bookvault's
# validator on 8.3 (2026-08-05): it expected 18 pages for a 20pp book, i.e.
# total minus 2.  Inside front and inside back covers are ordinary text pages.
#   20pp master -> 18pp text     16pp master -> 14pp text
COVER_FACES = 2

BLANK_COLUMNS = [
    "Publication Date",           # needs Lynden's on-sale date
    "EUR RRP", "USD RRP",         # GBP only for now
    "GBP %", "EUR %", "USD %",    # wholesale discount - only if selling trade
]

LEVEL_NAMES = {1: "Ditties", 2: "First Sounds", 3: "Special Friends",
               4: "Longer Sounds", 5: "New Spellings", 6: "Building Fluency",
               7: "Reading Together", 8: "Reading Champion"}


def page_counts() -> dict[str, int]:
    """Colour page count per book id, read from the real print masters."""
    out = {}
    for p in sorted(MASTERS.rglob("*.pdf")):
        if p.name.startswith("debug"):
            continue
        doc = fitz.open(p)
        out[p.stem.split(" ")[0].replace("_", ".")] = doc.page_count
        doc.close()
    return out


def short_description(title: str, level: int) -> str:
    return (f"A decodable phonics storybook for Level {level} "
            f"({LEVEL_NAMES.get(level, '')}). Every word uses only the sounds "
            f"taught by this level, so children read it for themselves.")


def long_description(story: dict, level: int) -> str:
    """Catalogue blurb built from the book's own data — opening line, the
    sounds it practises and the tricky words it introduces."""
    opening = (story.get("story_pages") or [{}])[0].get("text", "").strip()
    opening = " ".join(opening.split())
    if len(opening) > 150:
        opening = opening[:147].rsplit(" ", 1)[0] + "..."

    focus = [g for g in (story.get("focus_graphemes") or []) if g]
    tricky = [w for w in (story.get("tricky_words_used") or []) if w]

    parts = [
        f'"{opening}"' if opening else "",
        f"{story.get('book_title', '')} is a fully decodable storybook for "
        f"Level {level} ({LEVEL_NAMES.get(level, '')}) of the MyPhonicsBooks "
        f"reading journey — eight levels that take a child from their first "
        f"sounds to confident, fluent reading.",
    ]
    if focus:
        parts.append("This book practises the sounds "
                     + ", ".join(focus[:-1]) + (" and " if len(focus) > 1 else "")
                     + focus[-1] + ".")
    if tricky:
        parts.append("Tricky words that cannot yet be sounded out are shown "
                     "separately so a grown-up can read them together first: "
                     + ", ".join(tricky[:8]) + ".")
    parts.append(
        "Every word in the story uses only the sounds taught by this level or "
        "a listed tricky word, so the child reads it themselves rather than "
        "guessing from the pictures. Includes a sound chart, story words, "
        "talk-about questions and alien-word practice.")
    return " ".join(p for p in parts if p)


def main() -> int:
    template = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_TEMPLATE
    if not template.exists():
        sys.exit(f"Template not found: {template}\n"
                 f"Download it from Library > Bulk Upload Titles > Download Template")

    sys.path.insert(0, str(BASE_DIR / "scripts"))
    from isbn_barcodes import load_register
    from generate_pilot_books import get_pilot_stories, NEW_TO_OLD, LEVEL_KEYS

    register = load_register()
    pages = page_counts()
    stories = get_pilot_stories()

    missing = sorted(set(register) - set(pages))
    if missing:
        sys.exit(f"No print master for {missing} — build masters first so the "
                 f"page count is read from the real file, never assumed")

    wb = openpyxl.load_workbook(template)
    ws = wb["Titles"]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    # drop the example row; keep Sheet2's lookup tables intact
    ws.delete_rows(2, ws.max_row)

    book_ids = sorted(register, key=lambda k: [int(p) for p in k.split(".")])
    for n, book_id in enumerate(book_ids, start=2):
        entry = register[book_id]
        level = int(entry["level"])
        values = {
            "ISBN": entry["isbn"].replace("-", ""),
            "Title": entry["title"],
            "BindID": BIND_ID, "Binding": BIND_NAME,
            "Height": HEIGHT_MM, "Width": WIDTH_MM,
            "TextStockID": TEXT_STOCK_ID, "Text Stock": TEXT_STOCK,
            "CoverStockID": COVER_STOCK_ID, "Cover Stock": COVER_STOCK,
            "Colour Pages": pages[book_id] - COVER_FACES, "Mono Pages": 0,
            "LamID": LAM_ID, "Lamination": LAM_NAME,
            "Premium": "False", "Wholesale Type": "Not Wholesale",
            "LanguageID": LANGUAGE_ID, "Language": LANGUAGE,
            "CountryID": COUNTRY_ID, "Country Of Publication": COUNTRY,
            "AudID": AUD_ID, "Audience": AUDIENCE,
            "Imprint Name": IMPRINT, "Publisher": PUBLISHER,
            "Author Forename": AUTHOR_FORENAME,
            "Author Surname": AUTHOR_SURNAME,
            "City Of Publication": CITY,
            "BIC Code": BIC_CODE, "BISAC Code": BISAC_CODE,
            "Short Description": short_description(entry["title"], level),
            "Long Description": long_description(
                stories[LEVEL_KEYS[NEW_TO_OLD[book_id]]], level),
            "GBP RRP": rrp_gbp(level),
        }
        for header, value in values.items():
            ws.cell(n, col[header]).value = value

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / "titleUpload_MyPhonicsBooks.xlsx"
    wb.save(out)

    print(f"Wrote {out}")
    print(f"  {len(book_ids)} titles | {BIND_NAME} {HEIGHT_MM}x{WIDTH_MM}mm "
          f"| {TEXT_STOCK} | {COVER_STOCK} | {LAM_NAME}")
    counts = {}
    for b in book_ids:
        counts[pages[b] - COVER_FACES] = counts.get(pages[b] - COVER_FACES, 0) + 1
    print("  interior pages (cover excluded): "
          + ", ".join(f"{v} books at {k}pp" for k, v in sorted(counts.items())))
    print("\n  STILL BLANK — fill before uploading:")
    for c in BLANK_COLUMNS:
        print(f"    - {c}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

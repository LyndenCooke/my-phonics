"""Build the MPB Word Ledger — THE source-of-truth Excel for the whole scheme.

This workbook is the go-to for games, assessment, worksheets, word cards and
any future feature.  IMPORTANT: it is GENERATED from the canonical JSON data
(graphemes_by_level.json, tricky_words_by_level.json, shifty_sounds.json,
word banks, story dicts) — never hand-edit the xlsx; change the JSON and
re-run this script.

Sheets:
  1. "Sound Ledger"            — every grapheme in the 8-level ladder with
                                 its type, sound row, example, level colour
                                 and promotion/provenance notes, PLUS the
                                 final marking units (-le, -ed, -se, -ve)
                                 that aren't ladder graphemes but ARE single
                                 sound-button units.
  2. "Shifty Sounds"           — every card in shifty_sounds.json with its
                                 allowed-from level and band status
                                 (in band / promoted to ladder / removed).
  3. "Red Words (Tricky)"      — every tricky word, its level, and its true
                                 status from the tricky-word audit (true
                                 tricky vs graduates-at-L<n>).
  4. "Green Words by Sound"    — every decodable practice word in the
                                 curriculum data, grouped under the grapheme
                                 it practises (the latest-taught grapheme in
                                 its parse), with the level that grapheme
                                 lands at.  This is the source for printable
                                 word cards and on-site read-aloud practice
                                 beyond what fits in the books.
  5. "Green Words by Level"    — the same pool pivoted by level.
  6. "Held Back (review)"      — words excluded from the green pool pending
                                 a ruling / Shifty promotion.
  7. "Trap Words (books)"      — MANUAL-PASS WORKLIST: every story word in
                                 the 33 books that letter-decodes but hits a
                                 known phonetic trap at its book's level
                                 (final-y saying /ee/, soft c/g).  These are
                                 the 'happy'-class false passes the letter
                                 engine cannot judge — each needs Lynden's
                                 call: word-swap, tricky-list, or Watch Out.

Run:  py -3.12 scripts/build_word_ledger.py
Out:  output/worksheet_plan/MPB_WORD_LEDGER.xlsx
      output/worksheet_plan/green_words.json   (machine copy of the green
                                                pool for games / app code)
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE / "scripts"))

from audit_tricky_words import (  # noqa: E402
    PHONETICALLY_IRREGULAR, HONEST_PARSE_LEVEL, parses, cumulative_graphemes, load,
)

OUT = BASE / "output" / "worksheet_plan" / "MPB_WORD_LEDGER.xlsx"
JSON_OUT = BASE / "output" / "worksheet_plan" / "green_words.json"

LEVEL_NAMES = {1: "Ditties", 2: "First Sounds", 3: "Special Friends",
               4: "Longer Sounds", 5: "New Spellings", 6: "Building Fluency",
               7: "Reading Together", 8: "Reading Champion"}
LEVEL_HEX = {1: "E84B8A", 2: "F97066", 3: "F59E0B", 4: "22C55E",
             5: "3B82F6", 6: "6366F1", 7: "8B5CF6", 8: "14B8A6"}


def collect_words() -> set[str]:
    pool = set()

    def add(w):
        w = re.sub(r"[^a-z]", "", str(w).lower())
        if 2 <= len(w) <= 12:
            pool.add(w)

    # word banks (old 6-level files — still real word lists)
    for f in (BASE / "data" / "word_banks").glob("level_*_words.json"):
        data = json.load(open(f, encoding="utf-8"))
        def walk(x):
            if isinstance(x, str):
                add(x)
            elif isinstance(x, list):
                for i in x: walk(i)
            elif isinstance(x, dict):
                for v in x.values(): walk(v)
        walk(data)

    # spotlight words
    sp = json.load(open(BASE / "data" / "spotlight_words.json", encoding="utf-8"))
    for entry in sp.values():
        if not isinstance(entry, dict):
            continue
        for w in entry.get("words", []):
            add(w["word"] if isinstance(w, dict) else w)

    # story data word lists + story text
    sys.path.insert(0, str(BASE))
    from generate_pilot_books import get_pilot_stories  # type: ignore
    for story in get_pilot_stories().values():
        for k in ("story_words", "read_words", "writing_words"):
            for w in story.get(k, []):
                add(w)
        # Story text: only tokens that occur in lowercase somewhere — a token
        # seen ONLY capitalised is a proper noun (Yusuf, Mia), not a practice
        # word.
        for p in story.get("story_pages", []):
            for tok in re.findall(r"\b[a-z]+\b", p.get("text", "")):
                add(tok)

    return pool


def grapheme_level_map(g):
    m = {}
    for lv in range(1, 9):
        for gr in g.get(f"level_{lv}", {}).get("graphemes", []):
            m.setdefault(gr, lv)
            m.setdefault(gr.replace("-", ""), lv)
    return m


# Suffix graphemes only ever occur at the END of a word (optionally + s):
# 'ous' must not fire inside "houses".
SUFFIX_ONLY = {"ous", "able", "ible", "cious", "tious", "tion"}


def parse_units(word, graphemes):
    """Return the greedy parse as a list of units (None if no honest parse).

    Split digraphs are handled first: a word ending V-C-e where the V-e
    grapheme is taught parses as [...prefix..., V-e, C] (cake = c + a-e + k).
    """
    raw = set(graphemes)
    # The magic-e vowel may not itself follow another vowel ("choice" is
    # ch-oi-ce, not cho + i-e + c).
    m = re.match(r"^(.*?)([aiou]|e)([bcdfgklmnpstvz])e$", word)
    if (m and f"{m.group(2)}-e" in raw
            and (not m.group(1) or m.group(1)[-1] not in "aeiou")):
        prefix = parse_units(m.group(1), graphemes) if m.group(1) else []
        if prefix is not None:
            return prefix + [f"{m.group(2)}-e", m.group(3)]

    gs = sorted({x.replace("-", "") for x in graphemes}, key=len, reverse=True)
    units, i = [], 0
    while i < len(word):
        for cand in gs:
            if word.startswith(cand, i):
                if (len(cand) == 1 and cand in "aeiou"
                        and i + 1 < len(word) and word[i + 1] == cand):
                    return None
                if cand in SUFFIX_ONLY and word[i + len(cand):] not in ("", "s"):
                    continue
                units.append(cand)
                i += len(cand)
                break
        else:
            return None
    return units


# ── Sound Ledger helpers ─────────────────────────────────────────────────

VOWEL_LETTERS = set("aeiou")

# Final marking units: not ladder graphemes, but rendered as ONE sound-button
# unit in print + interactive (see PHONICS_PEDAGOGY.md "Marking grammar").
FINAL_UNITS = [
    ("-le", "final stable syllable", "purple, little, middle",
     "One under-line unit after a consonant; defers to able/ible."),
    ("-ed", "past-tense ending /d/ /t/ /id/", "turned, jumped, wanted",
     "One under-line unit; guards: len>=5, not a magic-e stem (liked)."),
    ("-se", "word-final /s/ or /z/, silent-e spelling", "purse, house, cheese",
     "Lynden 2026-07-12: ONE unit, silent e never gets its own dot. "
     "Only when the preceding vowel is consumed by a digraph or a consonant "
     "precedes the s — lone vowel before s = magic-e (close)."),
    ("-ve", "word-final /v/, silent-e spelling", "give, live, love",
     "Same rule as -se (no English word ends in v). Lone vowel before v = "
     "magic-e (wave, five). 've' must NEVER enter the ladder as a bare "
     "grapheme — tried and reverted 2026-07-12, it breaks wave/five."),
]

GRAPHEME_NOTES = {
    "wh": "Promoted from Shifty Sounds into the ladder 2026-07-12 (locked 07-03).",
    "ph": "Promoted from Shifty Sounds into the ladder 2026-07-12 (locked 07-03).",
    "sion": "In the L8 list but no L8 book teaches it yet — curriculum gap, do not use in activity words.",
}


def grapheme_type(gr: str, level_entry: dict) -> str:
    if "-" in gr:
        return "split digraph"
    if gr in level_entry.get("suffixes", []):
        return "suffix"
    if len(gr) == 1:
        return "single letter"
    if len(gr) == 2:
        return "vowel digraph" if gr[0] in VOWEL_LETTERS else "consonant digraph"
    if len(gr) == 3:
        return "trigraph"
    return "quadgraph+"


# ── Trap-word sweep (the 'happy' class) ──────────────────────────────────
# Words that LETTER-decode at their book's level but phonetically mislead —
# the false passes the letter engine cannot judge.  Output feeds the manual
# pass, it never auto-fixes.

def trap_reason(word: str, level: int) -> str | None:
    w = word.lower()
    # final -y saying /ee/ in a multi-syllable word (happy, funny, party).
    # y itself is taught at L2 but only as /y/ (yes); y=/ee/ has no ladder
    # slot (pronunciation promotion not implemented).
    if len(w) >= 4 and w.endswith("y") and w[-2] not in VOWEL_LETTERS \
            and sum(ch in VOWEL_LETTERS for ch in w) >= 1:
        return "final y says /ee/ — child reads /y/; needs swap, tricky-listing or Watch Out"
    # soft c / soft g before e, i, y — reads /s/ /j/, taught nowhere until
    # the (unimplemented) soft-c/soft-g promotion; L5.1 has a Watch Out.
    if re.search(r"c[eiy]", w):
        return "soft c says /s/ — child reads /k/; needs Watch Out or swap"
    # g before e/i/y is unpredictable in English (get/girl/gift are HARD g),
    # so exclude the common hard-g early-reader stems and flag the rest for
    # a human verdict rather than asserting.
    HARD_G = {"get", "gets", "got", "girl", "girls", "gift", "gifts", "give",
              "gives", "begin", "begins", "giggle", "giggles", "gear",
              "geese", "getting", "gig", "tiger", "together", "forget",
              "forgets", "target", "finger", "fingers", "longer", "stronger",
              "hunger", "anger", "burger", "burgers"}
    # doubled-g + suffix (hugged, jogging, foggy) is always HARD g — skip.
    if re.search(r"g[eiy]", w) and w not in HARD_G \
            and not re.search(r"gg(ed|ing|y|ier|iest)$", w):
        return "g before e/i/y — VERIFY: if it says /j/ (gem, magic, page) it needs Watch Out or swap"
    return None


def sweep_trap_words(t) -> list:
    sys.path.insert(0, str(BASE))
    from generate_pilot_books import get_pilot_stories, LEVEL_KEYS, NEW_TO_OLD
    tricky_all = {w.lower() for lv in range(1, 9)
                  for w in t.get(f"level_{lv}", {}).get("new_tricky_words", [])}
    stories = get_pilot_stories()
    rows, seen = [], set()
    for new_id in sorted(NEW_TO_OLD, key=lambda k: [int(p) for p in k.split(".")]):
        story = stories.get(LEVEL_KEYS.get(NEW_TO_OLD[new_id]))
        if not story:
            continue
        level = int(new_id.split(".")[0])
        author_tricky = {w.lower() for w in story.get("tricky_words_used", [])}
        pron_covered = set()
        for note in story.get("pronunciation_notes", []):
            for ex in note.get("examples", []):
                pron_covered.add(ex.split("→")[0].strip().lower())
        text = " ".join(p["text"] for p in story.get("story_pages", []))
        words = {tok.strip("'").lower() for tok in re.findall(r"[A-Za-z']+", text)}
        for k in ("story_words", "read_words", "writing_words"):
            words |= {str(w).lower() for w in story.get(k, [])}
        for w in sorted(words):
            if not w or w in tricky_all or w in author_tricky or w in pron_covered:
                continue
            reason = trap_reason(w, level)
            if reason and (new_id, w) not in seen:
                seen.add((new_id, w))
                rows.append((f"L{new_id}", story.get("book_title", "?"), w, reason))
    return rows


def main():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    g, t = load()
    glevel = grapheme_level_map(g)
    all_graphemes = cumulative_graphemes(g, 8)
    tricky_all = {w.lower() for lv in range(1, 9)
                  for w in t.get(f"level_{lv}", {}).get("new_tricky_words", [])}

    # ---- classify green words -------------------------------------------
    bare_graphemes = {x.replace("-", "") for x in all_graphemes}
    by_sound: dict[str, list] = {}
    review_ed: list = []
    for w in sorted(collect_words()):
        if (w in tricky_all or w in PHONETICALLY_IRREGULAR
                or w in HONEST_PARSE_LEVEL or w in bare_graphemes):
            continue
        units = parse_units(w, all_graphemes)
        if not units:
            continue
        # Past-tense -ed forms: 'picked' is /pikt/ — the -ed sounds aren't
        # taught, so these go to a review sheet, not the green pool.
        if (len(w) >= 5 and w.endswith("ed") and not w[-3] in "aeiou"
                and parse_units(w[:-2], all_graphemes)):
            review_ed.append((w, "-ed ending: /d/ /t/ /id/ not taught until the -ed promotion"))
            continue
        # Families that read wrong until their Shifty promotion lands:
        # soft c (face), soft g (age), -ve (give/live), y→i (cried), open-a
        # (famous, before-style open vowels are already tricky-listed).
        if re.search(r"c[eiy]", w) or re.search(r"g[eiy]", w):
            review_ed.append((w, "soft c/g: /s/ /j/ pending the soft-c / soft-g promotion"))
            continue
        if w.endswith("ive") or w.endswith("ved") or w in ("famous", "before"):
            review_ed.append((w, "needs an untaught sound (ve ending / open vowel)"))
            continue
        if w.endswith("ied") or w.endswith("ies"):
            review_ed.append((w, "y→i morphology: taught at L8 suffix work"))
            continue
        target = max(units, key=lambda u: glevel.get(u, 0))
        level = glevel.get(target, 0)
        by_sound.setdefault(target, []).append((w, level, "-".join(units)))

    wb = Workbook()

    # ---- Sheet: Sound Ledger --------------------------------------------
    ws0 = wb.active
    ws0.title = "Sound Ledger"
    ws0.append(["Level", "Level name", "Grapheme", "Type", "Example words", "Notes"])
    # Pull one example per grapheme from spotlight data when available.
    sp = json.load(open(BASE / "data" / "spotlight_words.json", encoding="utf-8"))
    def example_for(gr):
        entry = sp.get(gr) or sp.get(gr.replace("-", "_"))
        if isinstance(entry, dict):
            ws_ = entry.get("words", [])
            names = [w["word"] if isinstance(w, dict) else w for w in ws_[:3]]
            if names:
                return ", ".join(names)
        return ""
    for lv in range(1, 9):
        entry = g.get(f"level_{lv}", {})
        for gr in entry.get("graphemes", []):
            ws0.append([f"L{lv}", LEVEL_NAMES[lv], gr, grapheme_type(gr, entry),
                        example_for(gr), GRAPHEME_NOTES.get(gr, "")])
            ws0.cell(ws0.max_row, 3).fill = PatternFill("solid", fgColor=LEVEL_HEX[lv])
            ws0.cell(ws0.max_row, 3).font = Font(bold=True, color="FFFFFF")
        for sfx in entry.get("suffixes", []):
            if sfx in entry.get("graphemes", []):
                continue
            ws0.append([f"L{lv}", LEVEL_NAMES[lv], f"-{sfx}", "suffix (morphology)", "",
                        entry.get("morphology_note", "")[:80]])
        for pfx in entry.get("prefixes", []):
            ws0.append([f"L{lv}", LEVEL_NAMES[lv], f"{pfx}-", "prefix (morphology)", "", ""])
    for unit, sound, examples, note in FINAL_UNITS:
        ws0.append(["—", "Marking unit", unit, "final unit (one sound-button)",
                    examples, f"{sound}. {note}"])
        ws0.cell(ws0.max_row, 3).fill = PatternFill("solid", fgColor="475569")
        ws0.cell(ws0.max_row, 3).font = Font(bold=True, color="FFFFFF")
    for c in ws0[1]:
        c.font = Font(bold=True)
    ws0.column_dimensions["E"].width = 32
    ws0.column_dimensions["F"].width = 80

    # ---- Sheet: Shifty Sounds -------------------------------------------
    shifty = json.load(open(BASE / "data" / "shifty_sounds.json", encoding="utf-8"))
    promoted = {"wh", "ph"}
    ws_sh = wb.create_sheet("Shifty Sounds")
    ws_sh.append(["Grapheme", "Card type", "Sound", "Examples", "Allowed from", "Band status"])
    for card in shifty.get("alt_pronunciation_cards", []):
        gr = card["grapheme"]
        for i, pron in enumerate(card.get("pronunciations", [])):
            status = ("base sound (main ladder)" if i == 0
                      else "promoted to ladder 07-12" if gr in promoted
                      else "IN BAND (diamond-mark eligible)")
            ws_sh.append([gr, "alternative pronunciation", pron["sound"],
                          ", ".join(pron.get("examples", [])),
                          f"L{pron.get('allowed_from_level', '?')}", status])
    for card in shifty.get("new_spelling_cards", []):
        gr = card["grapheme"]
        status = ("promoted to ladder 07-12" if gr.lstrip("-") in promoted
                  else "IN BAND (never diamond-marked — alt spelling)")
        ws_sh.append([gr, "alternative spelling", card.get("sound", ""),
                      ", ".join(card.get("examples", [])),
                      f"L{card.get('allowed_from_level', '?')}", status])
    for c in ws_sh[1]:
        c.font = Font(bold=True)
    ws_sh.column_dimensions["D"].width = 42
    ws_sh.column_dimensions["F"].width = 44

    # ---- Sheet: Red words -----------------------------------------------
    ws = wb.create_sheet("Red Words (Tricky)")
    ws.append(["Level", "Level name", "Word", "Status"])
    for lv in range(1, 9):
        for word in t.get(f"level_{lv}", {}).get("new_tricky_words", []):
            clean = re.sub(r"[^a-z]", "", word.lower())
            if clean in HONEST_PARSE_LEVEL:
                status = f"graduates at L{HONEST_PARSE_LEVEL[clean]}"
            else:
                earliest = next((l for l in range(1, 9)
                                 if parses(clean, cumulative_graphemes(g, l))), None)
                if clean in PHONETICALLY_IRREGULAR or earliest is None:
                    status = "true tricky"
                else:
                    status = f"graduates at L{earliest}"
            ws.append([f"L{lv}", LEVEL_NAMES[lv], word, status])
    for c in ws[1]:
        c.font = Font(bold=True)

    # ---- Sheet 2: Green by sound ----------------------------------------
    ws2 = wb.create_sheet("Green Words by Sound")
    ws2.append(["Sound", "Taught at", "Level name", "Words (decodable practice)", "Count"])
    for gr in all_graphemes:
        key = gr
        entries = by_sound.get(key, [])
        words = ", ".join(w for w, _, _ in entries)
        row = [gr, f"L{glevel.get(key, '?')}", LEVEL_NAMES.get(glevel.get(key, 0), ""),
               words, len(entries)]
        ws2.append(row)
        fill = LEVEL_HEX.get(glevel.get(key, 0))
        if fill:
            ws2.cell(ws2.max_row, 1).fill = PatternFill("solid", fgColor=fill)
            ws2.cell(ws2.max_row, 1).font = Font(bold=True, color="FFFFFF")
    for c in ws2[1]:
        c.font = Font(bold=True)
    ws2.column_dimensions["D"].width = 110

    # ---- Sheet 3: Green by level ----------------------------------------
    ws3 = wb.create_sheet("Green Words by Level")
    ws3.append(["Level", "Level name", "Sound", "Words"])
    for lv in range(1, 9):
        for gr in g.get(f"level_{lv}", {}).get("graphemes", []):
            entries = by_sound.get(gr, [])
            ws3.append([f"L{lv}", LEVEL_NAMES[lv], gr,
                        ", ".join(w for w, _, _ in entries)])
    for c in ws3[1]:
        c.font = Font(bold=True)
    ws3.column_dimensions["D"].width = 110

    # ---- Sheet: words held back pending a ruling / promotion ------------
    ws4 = wb.create_sheet("Held Back (review)")
    ws4.append(["Word", "Why held back"])
    for w, reason in review_ed:
        ws4.append([w, reason])
    for c in ws4[1]:
        c.font = Font(bold=True)
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 70

    # ---- Sheet: trap words in the 33 books (manual-pass worklist) -------
    trap_rows = sweep_trap_words(t)
    ws5 = wb.create_sheet("Trap Words (books)")
    ws5.append(["Book", "Title", "Word", "Why it's a trap / what to decide"])
    for row in trap_rows:
        ws5.append(list(row))
    for c in ws5[1]:
        c.font = Font(bold=True)
    ws5.column_dimensions["B"].width = 30
    ws5.column_dimensions["D"].width = 90

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    # ---- JSON export of the green pool (for games / app code) -----------
    flat = [{"word": w, "sound": gr, "level": lvl, "units": units.split("-")}
            for gr, entries in by_sound.items()
            for w, lvl, units in entries]
    flat.sort(key=lambda e: (e["level"], e["sound"], e["word"]))
    json.dump(
        {"_note": ("GENERATED by scripts/build_word_ledger.py — never hand-edit. "
                   "Same pool as the ledger's Green Words sheets: tricky-listed, "
                   "-ed forms, soft c/g and other pre-promotion words are already "
                   "filtered out. sound = latest-taught grapheme in the parse; "
                   "level = where that grapheme lands; units = full sound-button "
                   "parse."),
         "words": flat},
        open(JSON_OUT, "w", encoding="utf-8"), indent=1, ensure_ascii=False)
    print(f"saved {JSON_OUT.relative_to(BASE)} ({len(flat)} words)")

    total = sum(len(v) for v in by_sound.values())
    print(f"saved {OUT.relative_to(BASE)}")
    print(f"green words: {total} across {len([k for k, v in by_sound.items() if v])} sounds; "
          f"red words: {len(tricky_all)}; trap words flagged: {len(trap_rows)}")


if __name__ == "__main__":
    main()
